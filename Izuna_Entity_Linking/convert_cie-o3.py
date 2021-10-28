import openpyxl
from pathlib import Path
import json

def readExcel(filename):
    xlsx_file = Path(filename + '.xlsx')
    wb_obj = openpyxl.load_workbook(xlsx_file)
    sheet = wb_obj.active

    return sheet

def createDictionary(sheet, choice_name = False):
    data = {}

    if choice_name == False:
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i == 0:
                data[row[0]] = []
                data[row[1]] = []
                data[row[2]] = []
            else:
                data['codigo'].append(row[0])
                data['descriptor completo'].append(row[1])
                data['descriptor corto'].append(row[2])
    else:
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i == 0:
                data[row[0]] = []
                data[row[1]] = []
            else:
                data['Behavior'].append(row[0])
                data['Canonical_name'].append(row[1])
    return data


def mergeSheets(datas, canonical_names):
    ind_x = 0
    for code in datas['codigo']:
        ind_y = 0
        for name in canonical_names['Behavior']:
            if code == name:
                datas['namos'][ind_x] = canonical_names['Canonical_name'][ind_y]
                continue
            ind_y += 1
        ind_x += 1
    return datas


def GenerateJSON(all_datas):   
    dict_datas = {}
    dict_txt = ''
    with open('cie_o3.json', 'w', encoding='utf-8') as json_file:
        for i in range(len(all_datas['codigo'])):
            namos = ''
            create_namos = str(all_datas['descriptor completo'][i]).split(' ')
            create_namos = [content for content in create_namos if not content in '']
            for n in create_namos:
                namos += str(n)[0]
            cortos = []
            if str(all_datas['descriptor corto'][i]).find(',') != -1:
                cortos = str(all_datas['descriptor corto'][i]).split(',')       
                cortos = [content.strip() for content in cortos if not content in '']
            else:
                cortos.append(str(all_datas['descriptor corto'][i]).strip())
            dict_datas = {"concept_id": all_datas['codigo'][i], "aliases": cortos, "canonical_name": namos, "definition": all_datas['descriptor completo'][i]}
            json.dump(dict_datas, json_file, ensure_ascii=False)
            json_file.write("\n")
    

sheets_1 = readExcel('CIEO31_table')
#sheets_2 = readExcel('icdo3_names')
datas = createDictionary(sheets_1, False)
#canonical_names = createDictionary(sheets_2, True)
#all_datas = mergeSheets(datas, canonical_names)
GenerateJSON(datas)
